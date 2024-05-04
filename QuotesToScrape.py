import openpyxl
import requests
import pandas
from bs4 import BeautifulSoup
import pandas as pd

#Open excel
FILE = 'test.xlsx'
wb = openpyxl.load_workbook(FILE)
ws = wb.active

#Connect to website
url = 'https://quotes.toscrape.com/'
response = requests.get(url)

#Parse the html of website
doc = BeautifulSoup(response.text,'html.parser')

#find all div with class 'quote'
div_tags = doc.find_all('div',class_='quote')

def get_quotes(div_tags):
    #Get all quotes on page
    quotes = []
    for tag in div_tags:
        quote = tag.find('span',class_='text').text
        quotes.append(quote)
    return quotes

def get_authors(div_tags):
    authors = []
    for tag in div_tags:
        span = tag.find('span',class_=None)
        author = span.find('small',class_='author').text
        authors.append(author)
    return authors

def get_quote_tag(div_tags):
    name_tags = []
    for tag in div_tags:
        name_tag = tag.find('div',class_='tags').meta['content']
        name_tags.append(name_tag)
    return name_tags

def get_author_links(div_tags):
    links = []
    for tag in div_tags:
        span = tag.find('span',class_=None)
        link = 'http://quotes.toscrape.com' + span.find('a')['href']
        links.append(link)
    return links

def convert_to_dict(quotes,authors,tags,links):
    dict = []
    for i in range(len(quotes)):
        dict.append({'Quote':quotes[i],'Author':authors[i],
                     'Tags':tags[i],'Biography link':links[i]},)
    return dict

def generate_list(div_tag):
    quotes = get_quotes(div_tags)
    authors = get_authors(div_tags)
    tags = get_quote_tag(div_tags)
    links = get_author_links(div_tags)

    dict = []
    for i in range(len(quotes)):
        dict.append({'Quote':quotes[i],'Author':authors[i],
                     'Tags':tags[i],'Biography link':links[i]},)
    return dict


dict = generate_list(div_tags)
headers = [ws.cell(row=1,column=i).value for i in range(1,5)]
for i in range(len(dict)):
    d = dict[i]
    for index,item in enumerate(headers,1):
        ws.cell(row=i+1,column = index,value = d[item])
wb.save('test.xlsx')
